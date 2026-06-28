"""Excel (.xlsx) extractor with per-sheet classification.

Uses openpyxl for cell-level metadata + pandas for tabular handling.
Each sheet is classified (HR / FUEL / WATER / WASTE / FINANCIAL / GHG / GENERIC)
based on column-header similarity to METRIC_REGISTRY aliases. Per-sheet handlers
emit RawField candidates so we can short-circuit the LLM for clean structured
data.
"""
from __future__ import annotations

import asyncio
import io
import re
from dataclasses import dataclass
from typing import Any, Callable, Optional

import openpyxl
import pandas as pd

from app.extractors.base import BaseExtractor, ExtractionContext, ExtractionResult
from app.models.internal import ChunkKind, DocumentChunk, RawField
from app.registry import METRIC_REGISTRY, find_by_alias
from app.utils.logging import get_logger
from app.utils.units import canonical_unit, parse_numeric

logger = get_logger("extractors.excel")


# ---------------------------------------------------------------------------
# Sheet type detection
# ---------------------------------------------------------------------------

SHEET_TYPE_KEYWORDS: dict[str, list[str]] = {
    "HR": ["employee", "emp id", "gender", "designation", "headcount", "joining date", "salary", "payroll"],
    "FUEL": ["diesel", "petrol", "lpg", "lng", "fuel", "hsd", "fuel oil", "litres consumed"],
    "WATER": ["water", "kl", "groundwater", "borewell", "discharge", "withdrawal", "effluent"],
    "WASTE": ["waste", "hazardous", "manifest", "e-waste", "plastic waste", "landfill", "incineration"],
    "FINANCIAL": ["turnover", "revenue", "capex", "opex", "ebitda", "profit", "balance sheet"],
    "GHG": ["co2", "tco2e", "scope 1", "scope 2", "scope 3", "emission factor"],
    "EHS": ["incident", "fatality", "ltifr", "trifr", "near miss", "safety"],
}


def classify_sheet(headers: list[str]) -> tuple[str, float]:
    hs = " ".join(h.lower() for h in headers if h)
    best_type, best_score = "GENERIC", 0.0
    for st, kws in SHEET_TYPE_KEYWORDS.items():
        hits = sum(1 for k in kws if k in hs)
        score = hits / max(len(kws), 1)
        if score > best_score:
            best_type, best_score = st, score
    return best_type, best_score


# ---------------------------------------------------------------------------
# Header → canonical metric mapping
# ---------------------------------------------------------------------------


def _map_headers_to_metrics(headers: list[str]) -> dict[int, str]:
    """Return {col_index: canonical_key} for headers that match a metric alias."""
    out: dict[int, str] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        key = find_by_alias(h)
        if key:
            out[i] = key
    return out


# ---------------------------------------------------------------------------
# Per-sheet handlers
# ---------------------------------------------------------------------------


@dataclass
class _SheetHandlerOutput:
    raw_fields: list[RawField]
    chunks: list[DocumentChunk]


def _handle_generic_table(
    sheet_name: str,
    df: pd.DataFrame,
    chunk_id_fn: Callable[[int], str],
) -> _SheetHandlerOutput:
    raw_fields: list[RawField] = []
    chunks: list[DocumentChunk] = []
    headers = [str(c) for c in df.columns]
    col_to_metric = _map_headers_to_metrics(headers)
    # Emit one chunk for the whole sheet as text — entity agent can re-scan.
    text = df.to_csv(index=False)
    chunks.append(
        DocumentChunk(
            chunk_id=chunk_id_fn(0),
            text=text,
            kind=ChunkKind.TABLE,
            sheet=sheet_name,
            table=[headers] + df.astype(str).values.tolist(),
            meta={"source": "excel", "rows": int(df.shape[0]), "cols": int(df.shape[1])},
        )
    )
    # Aggregate columns that mapped cleanly to a metric.
    for col_idx, key in col_to_metric.items():
        defn = METRIC_REGISTRY.get(key)
        if not defn:
            continue
        col_name = headers[col_idx]
        series = pd.to_numeric(df.iloc[:, col_idx], errors="coerce").dropna()
        if series.empty:
            continue
        total = float(series.sum())
        if total <= 0 and defn.get("value_constraints", {}).get("min", 0) >= 0:
            continue
        raw_fields.append(
            RawField(
                canonical_key=key,
                raw_label=col_name,
                raw_value=str(total),
                value_num=total,
                unit=defn.get("unit"),
                sheet=sheet_name,
                cell=f"{sheet_name}!{_col_letter(col_idx)}",
                source="sheet_handler",
                notes=f"sum of column '{col_name}' ({len(series)} values)",
            )
        )
    return _SheetHandlerOutput(raw_fields=raw_fields, chunks=chunks)


def _handle_hr_sheet(
    sheet_name: str,
    df: pd.DataFrame,
    chunk_id_fn: Callable[[int], str],
) -> _SheetHandlerOutput:
    out = _handle_generic_table(sheet_name, df, chunk_id_fn)
    raw_fields = out.raw_fields

    cols = {c.lower().strip(): c for c in df.columns.astype(str)}
    total = len(df)
    raw_fields.append(
        RawField(
            canonical_key="employee_count_total",
            raw_label="rowcount",
            raw_value=str(total),
            value_num=float(total),
            unit="count",
            sheet=sheet_name,
            source="sheet_handler",
            notes="row count of HR sheet",
        )
    )

    gender_col = next((cols[c] for c in cols if "gender" in c or "sex" in c), None)
    if gender_col:
        series = df[gender_col].astype(str).str.lower().str.strip()
        male = int(series.isin(["m", "male", "man"]).sum())
        female = int(series.isin(["f", "female", "woman"]).sum())
        lgbtq = int(series.isin(["lgbtq", "lgbtqia", "other", "transgender", "non-binary", "nb"]).sum())
        raw_fields.extend(
            [
                RawField(
                    canonical_key="employee_count_male",
                    raw_label=gender_col,
                    raw_value=str(male),
                    value_num=float(male),
                    unit="count",
                    sheet=sheet_name,
                    source="sheet_handler",
                ),
                RawField(
                    canonical_key="employee_count_female",
                    raw_label=gender_col,
                    raw_value=str(female),
                    value_num=float(female),
                    unit="count",
                    sheet=sheet_name,
                    source="sheet_handler",
                ),
            ]
        )
        if lgbtq:
            raw_fields.append(
                RawField(
                    canonical_key="employee_count_lgbtq",
                    raw_label=gender_col,
                    raw_value=str(lgbtq),
                    value_num=float(lgbtq),
                    unit="count",
                    sheet=sheet_name,
                    source="sheet_handler",
                )
            )

    return _SheetHandlerOutput(raw_fields=raw_fields, chunks=out.chunks)


SHEET_HANDLERS: dict[str, Callable[[str, pd.DataFrame, Callable[[int], str]], _SheetHandlerOutput]] = {
    "HR": _handle_hr_sheet,
    # FUEL/WATER/WASTE/FINANCIAL/GHG use generic + metric-alias matching, which
    # already extracts column sums.
    "FUEL": _handle_generic_table,
    "WATER": _handle_generic_table,
    "WASTE": _handle_generic_table,
    "FINANCIAL": _handle_generic_table,
    "GHG": _handle_generic_table,
    "EHS": _handle_generic_table,
    "GENERIC": _handle_generic_table,
}


# ---------------------------------------------------------------------------
# Extractor
# ---------------------------------------------------------------------------


class ExcelExtractor(BaseExtractor):
    name = "excel"

    async def extract(self, ctx: ExtractionContext) -> ExtractionResult:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, self._sync_extract, ctx)

    def _sync_extract(self, ctx: ExtractionContext) -> ExtractionResult:
        result = ExtractionResult()
        preview_parts: list[str] = []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(ctx.file_bytes), data_only=True, read_only=True)
        except Exception as e:  # noqa: BLE001
            logger.error("excel.load_failed", err=str(e))
            result.notes.append(f"excel load failed: {e}")
            return result

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            try:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                # Heuristic: header row = first non-empty row
                header_row_idx = 0
                for i, r in enumerate(rows):
                    if any(c is not None and str(c).strip() for c in r):
                        header_row_idx = i
                        break
                headers = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
                data_rows = rows[header_row_idx + 1 :]
                # Pad short rows
                max_cols = max(len(headers), max((len(r) for r in data_rows), default=0))
                headers += [""] * (max_cols - len(headers))
                data_rows = [list(r) + [None] * (max_cols - len(r)) for r in data_rows]
                df = pd.DataFrame(data_rows, columns=headers)

                stype, score = classify_sheet(headers)
                logger.info("excel.sheet_classified", sheet=sheet_name, type=stype, score=round(score, 3))

                def chunk_id_fn(i: int, _s=sheet_idx) -> str:
                    return self._chunk_id("xls", _s + 1, i + 1)

                handler = SHEET_HANDLERS.get(stype, _handle_generic_table)
                out = handler(sheet_name, df, chunk_id_fn)
                result.chunks.extend(out.chunks)
                result.raw_fields.extend(out.raw_fields)

                if len(preview_parts) < 3:
                    head = df.head(5).to_csv(index=False)
                    preview_parts.append(f"# Sheet: {sheet_name}\n{head}")
            except Exception as e:  # noqa: BLE001
                logger.warning("excel.sheet_failed", sheet=sheet_name, err=str(e))
                result.notes.append(f"sheet '{sheet_name}' failed: {e}")

        wb.close()
        result.page_count = len(wb.sheetnames)
        result.text_preview = ("\n\n".join(preview_parts))[:2000]
        return result


def _col_letter(col_idx: int) -> str:
    """0-based index → Excel column letter."""
    s = ""
    n = col_idx
    while True:
        s = chr(65 + (n % 26)) + s
        n = n // 26 - 1
        if n < 0:
            break
    return s
