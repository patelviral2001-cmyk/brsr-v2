"""Shared pytest fixtures.

These fixtures mock all external services (LLMs, S3, Qdrant, Redis) so tests
can run hermetically.
"""
from __future__ import annotations

import io
from typing import Any

import pytest

from app.config import TaskType
from app.llm.router import LLMCall, LLMResult


# ---------------------------------------------------------------------------
# Mock LLM router
# ---------------------------------------------------------------------------


class FakeRouter:
    """Drop-in replacement for LLMRouter.

    Mirrors both the legacy ``chat()`` signature and the new ``route()``
    signature so tests can patch ``get_router`` and exercise either entry
    point. Canned responses are queued per ``TaskType`` and consumed FIFO.
    """

    def __init__(self) -> None:
        self.calls: list[dict[str, Any]] = []
        # task -> list of canned dicts (consumed FIFO)
        self.queue: dict[TaskType, list[dict[str, Any]]] = {t: [] for t in TaskType}

    def enqueue(
        self,
        task: TaskType,
        parsed: dict[str, Any],
        text: str = "",
        *,
        prompt_tokens: int = 0,
        completion_tokens: int = 0,
    ) -> None:
        self.queue[task].append(
            {
                "parsed": parsed,
                "text": text,
                "prompt_tokens": prompt_tokens,
                "completion_tokens": completion_tokens,
            }
        )

    # ------------------------------------------------------------------
    # Legacy entrypoint (the one all agents use today)
    # ------------------------------------------------------------------
    async def chat(
        self,
        task: TaskType,
        messages: list[dict[str, Any]],
        *,
        prompt_version: str = "",
        tenant_id: str = "",
        json_mode: bool = False,
        response_schema: dict[str, Any] | None = None,
        response_format: dict[str, Any] | None = None,
        tools: list[dict[str, Any]] | None = None,
        temperature: float | None = None,
        max_tokens: int | None = None,
        extra_meta: dict[str, Any] | None = None,
    ) -> LLMResult:
        self.calls.append(
            {
                "task": task,
                "messages": messages,
                "prompt_version": prompt_version,
                "response_format": response_format,
                "response_schema": response_schema,
                "tools": tools,
                "extra_meta": extra_meta,
            }
        )
        item = self._pop(task)
        text = item.get("text", "") or ""
        parsed = item.get("parsed")
        return LLMResult(
            text=text,
            parsed=parsed,
            raw={},
            call=LLMCall(
                task=task.value,
                model="mock",
                prompt_version=prompt_version,
                input_tokens=int(item.get("prompt_tokens", 0) or 0),
                output_tokens=int(item.get("completion_tokens", 0) or 0),
            ),
        )

    # ------------------------------------------------------------------
    # New canonical entrypoint
    # ------------------------------------------------------------------
    async def route(
        self,
        task: TaskType,
        messages: list[dict[str, Any]],
        *,
        tools: list[dict[str, Any]] | None = None,
        response_format: dict[str, Any] | None = None,
        tool_choice: Any | None = None,
        temperature: float | None = None,
        max_tokens: int | None = None,
        prompt_version: str = "",
        tenant_id: str = "",
        extra_meta: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        self.calls.append(
            {
                "task": task,
                "messages": messages,
                "prompt_version": prompt_version,
                "response_format": response_format,
                "tools": tools,
                "extra_meta": extra_meta,
            }
        )
        item = self._pop(task)
        text = item.get("text", "") or ""
        parsed = item.get("parsed")
        call = LLMCall(
            task=task.value,
            model="mock",
            prompt_version=prompt_version,
            input_tokens=int(item.get("prompt_tokens", 0) or 0),
            output_tokens=int(item.get("completion_tokens", 0) or 0),
        )
        return {
            "content": text,
            "tool_calls": [],
            "parsed": parsed,
            "model": "mock",
            "prompt_tokens": call.input_tokens,
            "completion_tokens": call.output_tokens,
            "latency_ms": 0,
            "cost": 0.0,
            "raw": {},
            "call": call,
        }

    def _pop(self, task: TaskType) -> dict[str, Any]:
        q = self.queue.get(task, [])
        if q:
            return q.pop(0)
        return {"parsed": {}, "text": ""}


@pytest.fixture
def fake_router(monkeypatch: pytest.MonkeyPatch) -> FakeRouter:
    router = FakeRouter()
    # Patch the singleton accessor
    monkeypatch.setattr("app.llm.router.get_router", lambda: router)
    monkeypatch.setattr("app.agents.document_classifier.get_router", lambda: router)
    monkeypatch.setattr("app.agents.entity_extraction_agent.get_router", lambda: router)
    monkeypatch.setattr("app.agents.validation_agent.get_router", lambda: router)
    return router


# ---------------------------------------------------------------------------
# Sample PDF / XLSX bytes
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_pdf_bytes() -> bytes:
    """Build a tiny PDF in-memory using PyMuPDF."""
    import fitz

    doc = fitz.open()
    page = doc.new_page()
    page.insert_text(
        (72, 72),
        "Acme Corp\nFY 2024-25\nTotal Electricity Consumption: 12,345 kWh\n"
        "Renewable Electricity: 3,200 kWh\nDiesel consumed: 4,500 litres\n",
        fontsize=11,
    )
    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_hr_bytes() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Headcount"
    ws.append(["Emp ID", "Name", "Gender", "Designation", "Joining Date"])
    rows = [
        ("E001", "A", "M", "Engineer", "2020-01-01"),
        ("E002", "B", "F", "Manager", "2019-05-12"),
        ("E003", "C", "M", "Engineer", "2021-09-15"),
        ("E004", "D", "F", "Engineer", "2022-02-01"),
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_fuel_bytes() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fuel"
    ws.append(["Month", "Diesel (litres)", "Petrol (litres)"])
    for m, d, p in [("Apr", 1000, 200), ("May", 1100, 220), ("Jun", 950, 180)]:
        ws.append([m, d, p])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_water_bytes() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Water"
    ws.append(["Month", "Groundwater (kL)", "Third Party (kL)", "Discharge (kL)"])
    for m, gw, tp, dc in [("Apr", 500, 100, 400), ("May", 480, 120, 380)]:
        ws.append([m, gw, tp, dc])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def sample_xlsx_waste_bytes() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Waste"
    ws.append(["Month", "Hazardous Waste (kg)", "Plastic Waste (kg)", "E-Waste (kg)"])
    for r in [("Apr", 100, 30, 5), ("May", 120, 25, 10)]:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Patch S3 download to return a provided byte fixture
# ---------------------------------------------------------------------------


@pytest.fixture
def patch_s3(monkeypatch: pytest.MonkeyPatch):
    state: dict[str, bytes] = {}

    async def fake_download(url: str) -> tuple[bytes, str]:
        if url not in state:
            raise RuntimeError(f"unmocked URL: {url}")
        data = state[url]
        return data, url.rsplit("/", 1)[-1] or "file.bin"

    monkeypatch.setattr("app.utils.s3.download_to_bytes", fake_download)
    monkeypatch.setattr("app.orchestrator.document_orchestrator.download_to_bytes", fake_download)
    return state


@pytest.fixture
def patch_rag(monkeypatch: pytest.MonkeyPatch):
    """No-op the RAG indexer so tests don't try to hit Qdrant."""
    called: list[dict[str, Any]] = []

    async def fake_index(**kwargs):
        called.append(kwargs)
        return 0

    monkeypatch.setattr(
        "app.rag.indexer.RagIndexer.index_document",
        fake_index,
    )
    return called
